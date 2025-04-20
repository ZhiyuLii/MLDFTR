import pyscf
from pyscf import gto
from pyscf import dft
from pyscf import scf
import numpy as np
import torch
from snn_3h import SNN
from snn_3hR import SNN2
import pandas as pd
import openpyxl
from ACONF import systems, reactions  # change the set you wanna test
from normal_coef import get_normal_coefficient
p = 40
f = 27000
fppath = 'S'+ str(p)+'-'+str(f)+'.npz'
pfname = str(p)+'-'+str(f)
SNc = get_normal_coefficient(p,f)

def getAtom(atom):
    atm = ''
    length = len(systems[atom]['atoms'])
    for i in range(length):
        atm += systems[atom]['atoms'][i]
        atm += ' '
        for j in range(len(systems[atom]['coords'][i])):
            atm += str(systems[atom]['coords'][i][j])
            atm += ' '
        if i != length-1:
            atm += ';'

    charge = systems[atom]['charge']
    spin = systems[atom]['spin']

    return atm, spin, charge




input_dim_nn1 = 3  # rho zeta  r_s
input_dim_nn2 = 4

output_dim = 1
depth = 4
lamda = 1e-5
beta = 1.5
use_cuda = False
device = torch.device("cuda" if use_cuda else "cpu")
scale_factor = 0.001
iterindex = 0
new_index = 0
scfloss = []
hidden = [30, 30, 30]

s_nn = SNN(input_dim_nn1, output_dim, hidden, lamda, beta, use_cuda)
model_path_nn1 = "nn1.pth"
s_nn.load_state_dict(torch.load(model_path_nn1))
s_nn.to(device)
nn2 = SNN2(input_dim_nn2, output_dim, hidden, lamda, beta, use_cuda)
model_path_nn2 = "ML-B3LYP-Ra.pth.pth"   #更改
nn2.load_state_dict(torch.load(model_path_nn2))
nn2.to(device)

def eval_xc_ml(xc_code, rho, spin, relativity=0, deriv=1, verbose=None):
    if spin == 0:
        rho0, dx, dy, dz = rho[:4]
        gamma1 = gamma2 = gamma12 = (dx ** 2 + dy ** 2 + dz ** 2) * 0.25 + 1e-10
        rho01 = rho02 = rho0 * 0.5
    else:
        rho1 = rho[0]
        rho2 = rho[1]
        rho01, dx1, dy1, dz1 = rho1[:4]
        rho02, dx2, dy2, dz2 = rho2[:4]
        gamma1 = dx1 ** 2 + dy1 ** 2 + dz1 ** 2 + 1e-10
        gamma2 = dx2 ** 2 + dy2 ** 2 + dz2 ** 2 + 1e-10
        gamma12 = dx1 * dx2 + dy1 * dy2 + dz1 * dz2 + 1e-10

    rhos = rho01 + rho02  # [grids,][0]=grids
    N = rhos.shape[0]  # list[0]=grids
    ml_in_ = np.concatenate((rho01.reshape((-1, 1)), rho02.reshape((-1, 1)), gamma1.reshape((-1, 1)),
                             gamma2.reshape((-1, 1)), gamma12.reshape((-1, 1))),
                            axis=1)
    ml_in = torch.Tensor(ml_in_)
    ml_in.requires_grad = True
    num_r = ml_in.shape[0]
    dim_in = ml_in.shape[1]
    exc_ml_out = s_nn(ml_in, is_training_data=False)
    ml_exc = exc_ml_out.detach().numpy()
    uni = np.power(rhos, 1 / 3) * 0.75 * np.power(3 / np.pi, 1 / 3)  # (grids,)
    ml_exc = ml_exc * (uni.reshape(N, 1))
    exc_ml = torch.dot(exc_ml_out[:, 0],
                       torch.pow((ml_in[:, 0] + ml_in[:, 1]), 4 / 3) * 0.75 * np.power(3 / np.pi, 1 / 3))
    exc_ml.backward()
    grad = ml_in.grad.data.numpy()  # grid[:,5] for v_r_rho should be held by artificial action as same as 6 and 7

    if spin != 0:
        vrho_ml = np.hstack(((grad[:, 0]).reshape((-1, 1)), (grad[:, 1]).reshape((-1, 1))))
        vgamma_ml = np.hstack((grad[:, 2].reshape((-1, 1)), grad[:, 4].reshape((-1, 1)), grad[:, 3].reshape((-1, 1))))
        vlapl = np.zeros((N, 2))
        vtau = np.zeros((N, 2))
    else:
        vrho_ml = (grad[:, 0] + grad[:, 1]) * 0.5
        vgamma_ml = (grad[:, 2] + grad[:, 3] + grad[:, 4]) * 0.25
        vlapl = np.zeros(N)
        vtau = np.zeros(N)
    b3lyp_xc = dft.libxc.eval_xc('B3LYP', rho, spin, relativity, deriv, verbose)
    b3lyp_exc = np.array(b3lyp_xc[0])
    b3lyp_vrho = np.array(b3lyp_xc[1][0])
    b3lyp_vgamma = np.array(b3lyp_xc[1][1])
    exc = (b3lyp_exc.reshape(-1, 1) + ml_exc).reshape(-1)
    vrho = b3lyp_vrho + vrho_ml
    vgamma = b3lyp_vgamma + vgamma_ml
    vxc = (vrho, vgamma, vlapl, vtau)
    fxc = None  # 2nd order functional derivative
    kxc = None  # 3rd order functional derivative
    return exc, vxc, fxc, kxc



def ml_in(rho,n):  # creator of input_tensor_rhop   将rho转化成input的tensor形式，实际上会组装成一维的tensor，在网络里会继续变形，变成[grids,3]的形状。3代表着 映射的三个参数
    if n==0:
        rho0, dx, dy, dz = rho[:4]
        gamma1 = gamma2 = gamma12 = (dx ** 2 + dy ** 2 + dz ** 2) * 0.25 + 1e-10
        rho01 = rho02 = rho0 * 0.5
    else:
        rho1 = rho[0]
        rho2 = rho[1]
        rho01, dx1, dy1, dz1 = rho1[:4]
        rho02, dx2, dy2, dz2 = rho2[:4]
        gamma1 = dx1 ** 2 + dy1 ** 2 + dz1 ** 2 + 1e-10
        gamma2 = dx2 ** 2 + dy2 ** 2 + dz2 ** 2 + 1e-10
        gamma12 = dx1 * dx2 + dy1 * dy2 + dz1 * dz2 + 1e-10
    ml_in_ = np.concatenate((rho01.reshape((-1, 1)), rho02.reshape((-1, 1)), gamma1.reshape((-1, 1)),
                             gamma2.reshape((-1, 1)), gamma12.reshape((-1, 1))), axis=1)
    return ml_in_  # , num_r, dim_in



def ml_inR(rho, weight, coords, n):
    if n==0:
        rho0, dx, dy, dz = rho[:4]
        gamma1 = gamma2 = gamma12 = (dx ** 2 + dy ** 2 + dz ** 2) * 0.25 + 1e-10
        rho01 = rho02 = rho0 * 0.5
    else:
        rho1 = rho[0]
        rho2 = rho[1]
        rho01, dx1, dy1, dz1 = rho1[:4]
        rho02, dx2, dy2, dz2 = rho2[:4]
        gamma1 = dx1 ** 2 + dy1 ** 2 + dz1 ** 2 + 1e-10
        gamma2 = dx2 ** 2 + dy2 ** 2 + dz2 ** 2 + 1e-10
        gamma12 = dx1 * dx2 + dy1 * dy2 + dz1 * dz2 + 1e-10
    rhos = rho01 + rho02  # [grids,][0]=grids
    S = np.load(fppath)       # take the selected frequency points
    k_m = S['k']
    f_m_r = S['fr']
    f_m_i = S['fi']
    coords1 = coords - (p/2)
    rw = np.multiply(rhos, weight)
    R1 = 0
    for i in range(0,f, 300):
        a = int(i)
        if i + 300 <= f:
            b = f
        else:
            b = int(i + 300)
        kr = np.einsum('ik,jk->ij', k_m[a:b], coords1)
        fsin = np.einsum('j, ij->i', rw, np.sin(kr))
        fcos = np.einsum('j, ij->i', rw, np.cos(kr))  # (Num,)
        kr = 0
        kr1 = np.einsum('ik, jk->ij', k_m[a:b], coords)
        frfs = np.multiply(fsin, f_m_r[a:b])
        frfc = np.multiply(fcos, f_m_r[a:b])
        fifs = np.multiply(fsin, f_m_i[a:b])
        fifc = np.multiply(fcos, f_m_i[a:b])
        R1 += np.einsum('i,ij->j', (frfc + fifs), np.cos(kr1))
        R1 += np.einsum('i,ij->j', (frfs - fifc), np.sin(kr1))
        kr1 =  frfs = frfc = fifs = fifc =0
    R1 *= 1 / 64 / 64 / 64 * SNc

    ml_in_ = np.concatenate((rho01.reshape((-1, 1)), rho02.reshape((-1, 1)), gamma1.reshape((-1, 1)),
                             gamma2.reshape((-1, 1)), gamma12.reshape((-1, 1)),
                             R1.reshape((-1, 1))
                             ),
                            axis=1)  # 这里输入了十个R，该张量在用作测试时，仅仅使用了[5:]之后的张量
    R1 = kr = frfs = frfc = fifs = fifc = 0
    return ml_in_  # , num_r, dim_in


s_nn.eval()
nn2.eval()
energy = dict()
for item in systems:
    # 这里的items是字符串格式， 可以用来做字典的key, 可以给计算的结果命名
    print(item)
    mole_info = getAtom(item)

    mole = pyscf.M(
        atom=mole_info[0],
        spin=mole_info[1],
        charge=mole_info[2],
        basis="def2-tzvpd",
        verbose=1 )

    if mole_info[1] == 0:
        mlpbe = dft.RKS(mole)
        mlpbe = mlpbe.define_xc_(eval_xc_ml, 'GGA', hyb=0.2, rsh=[0.0, 0.0, 0.2])
        mlpbe.grids.level = 5
        mlpbe.max_cycle = 50
        A = mlpbe.kernel()
        print('calculation is down')
        mycoords_ = mlpbe.grids.coords
        myweight_ = mlpbe.grids.weights
        mydm_ = mlpbe.make_rdm1()
        myao = dft.numint.eval_ao(mole, mlpbe.grids.coords, deriv=1)
        rho_ = dft.numint.eval_rho(mole, myao, mydm_, xctype='GGA')
        rho = ml_inR(rho_, myweight_, mycoords_, 0)
        myinput = torch.Tensor(rho)
        output = nn2(myinput, is_training_data=False)
        uni = torch.pow(myinput[:, 0] + myinput[:, 1], 1 / 3) * 0.75 * np.power(3 / np.pi, 1 / 3)
        output *= uni.unsqueeze(1)
        B = ((output * torch.tensor(myweight_).unsqueeze(1) * torch.tensor(rho_[0]).unsqueeze(1)).real).sum()
        C = (A + B.data.numpy()) * 627.5095
        B3 = dft.RKS(mole)
        B3.xc = 'b3lyp'
        B3.grids.level = 5
        B3.max_cycle = 50
        D = 0
        #D = B3.kernel()
    else:
        mlpbe = dft.UKS(mole)
        mlpbe = mlpbe.define_xc_(eval_xc_ml, 'GGA', hyb=0.2, rsh=[0.0, 0.0, 0.2])
        mlpbe.grids.level = 5
        mlpbe.max_cycle = 50
        A = mlpbe.kernel()
        print('calculation is down')
        mycoords_ = mlpbe.grids.coords
        myweight_ = mlpbe.grids.weights
        myao = dft.numint.eval_ao(mole, mlpbe.grids.coords, deriv=1)
        mydm_ = mlpbe.make_rdm1()
        rhoa = dft.numint.eval_rho(mole, myao, mydm_[0], xctype='GGA')  # ???这里应该有问题。
        rhob = dft.numint.eval_rho(mole, myao, mydm_[1], xctype='GGA')
        rho_ = (rhoa, rhob)
        rho = ml_inR(rho_, myweight_, mycoords_, 1)
        myinput = torch.Tensor(rho)
        output = nn2(myinput, is_training_data=False)
        uni = torch.pow(myinput[:, 0] + myinput[:, 1], 1 / 3) * 0.75 * np.power(3 / np.pi, 1 / 3)
        output *= uni.unsqueeze(1)
        B = (output * torch.tensor(myweight_).unsqueeze(1) * torch.tensor(rho_[0][0]).unsqueeze(
            1) + output * torch.tensor(
            myweight_).unsqueeze(1) * torch.Tensor(rho_[1][0]).unsqueeze(1)).sum()
        C = (A + B.data.numpy()) * 627.5095
        B3 = dft.UKS(mole)
        B3.xc = 'b3lyp'
        B3.grids.level = 5
        B3.max_cycle = 50
        D = 0
        #D = B3.kernel()
    result = [A, B, C, D]
    energy[item] = result

output = "aconf_Ra.xlsx"
wb = openpyxl.Workbook(output)
wb.save(output)
wb_read = openpyxl.load_workbook(output)
ws = wb_read.active
count = 2

# set table's title
ws.cell(1, 1, "Reaction")
ws.cell(1, 2, "Reference")
ws.cell(1, 3, "ML-DFT-nn1")
ws.cell(1, 4, "B3LYP")
ws.cell(1, 5, "ML-DFT-Ra")
ws.cell(1, 6, "nn1_Error")
ws.cell(1, 7, "B3LYP_Error")
ws.cell(1, 8, "Ra_Error")

for reaction in reactions:  #reactions是一个列表
    ws.cell(count, 1, str(reaction['systems']))
    atom_nums = len(reaction['systems']) #统计反应中物种的个数，例如这里基本为2
    ml_res = 0
    b3lyp_res = 0
    R_res = 0
    for i in range(atom_nums):
        ml_res += energy[reaction['systems'][i]][0] * int(reaction['stoichiometry'][i])
        b3lyp_res += energy[reaction['systems'][i]][3] * int(reaction['stoichiometry'][i])
        R_res += energy[reaction['systems'][i]][2] * int(reaction['stoichiometry'][i])
    # 参考能量
    ref = reaction['reference']
    ws.cell(count, 2, ref)

    # 能量值比较
    ws.cell(count, 3, ml_res * 627.5095)
    ws.cell(count, 4, b3lyp_res * 627.509)
    ws.cell(count, 5, R_res)
    # 误差分析
    ws.cell(count, 6, abs((ml_res * 627.509 - ref)))
    ws.cell(count, 7, abs((b3lyp_res * 627.509 - ref)))
    ws.cell(count, 8, abs((R_res  - ref)))
    count += 1
    wb_read.save(output)

# pandas mae
data = pd.read_excel(output)
ml_mae = data['nn1_Error'].mean(axis=0)
b3lyp_mae = data['B3LYP_Error'].mean(axis=0)
R_mae = data['Ra_Error'].mean(axis=0)
ws.cell(count, 6, ml_mae)
ws.cell(count, 7, b3lyp_mae)
ws.cell(count, 8, R_mae)
wb_read.save(output)







